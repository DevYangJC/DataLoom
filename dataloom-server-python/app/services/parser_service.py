import json
import datetime
import logging
from io import BytesIO
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.config import CHUNK_SIZE
from app.models import ExcelDocument, ExcelSheet, ExcelSheetChunk

logger = logging.getLogger(__name__)

def build_cell_value(val_cell, form_cell) -> Optional[Dict[str, Any]]:
    """
    Builds the Luckysheet-compatible cell value structure.
    val_cell is from the workbook opened with data_only=True.
    form_cell is from the workbook opened with data_only=False.
    """
    val = val_cell.value
    form = form_cell.value

    if val is None and form is None:
        return None

    v_dict = {}
    ct = {"fa": "General"}

    # Detect formula
    is_formula = False
    formula_str = None
    if isinstance(form, str) and form.startswith("="):
        is_formula = True
        formula_str = form

    # Determine type of the value
    # If formula, we want the evaluated value from val_cell. If not, from form_cell.
    actual_val = val if is_formula else form

    if actual_val is None:
        if is_formula:
            v_dict["f"] = formula_str
            v_dict["v"] = ""
            v_dict["m"] = ""
            ct["t"] = "s"
            v_dict["ct"] = ct
            return v_dict
        return None

    if isinstance(actual_val, bool):
        v_dict["v"] = actual_val
        v_dict["m"] = "TRUE" if actual_val else "FALSE"
        ct["t"] = "b"
    elif isinstance(actual_val, (int, float)):
        v_dict["v"] = actual_val
        # Format number as string
        if actual_val == int(actual_val):
            v_dict["m"] = str(int(actual_val))
        else:
            v_dict["m"] = str(actual_val)
        ct["t"] = "n"
    elif isinstance(actual_val, (datetime.datetime, datetime.date)):
        date_str = actual_val.strftime("%Y-%m-%d")
        v_dict["v"] = date_str
        v_dict["m"] = date_str
        ct["fa"] = "yyyy-MM-dd"
        ct["t"] = "d"
    else:
        # String type
        val_str = str(actual_val)
        if not val_str:
            return None
        v_dict["v"] = val_str
        v_dict["m"] = val_str
        ct["t"] = "s"

    if is_formula:
        v_dict["f"] = formula_str

    v_dict["ct"] = ct
    return v_dict


def get_column_width_pixels(sheet, col_idx: int) -> int:
    col_letter = get_column_letter(col_idx + 1)
    width = sheet.column_dimensions[col_letter].width
    if width is None:
        return 72
    return max(int(width * 8), 72)


def get_row_height_pixels(sheet, row_idx: int) -> Optional[int]:
    height = sheet.row_dimensions[row_idx + 1].height
    if height is None:
        return None
    return int(round(height / 0.75))


def build_merge_config(sheet) -> Dict[str, Any]:
    merge_config = {}
    for merged_range in sheet.merged_cells.ranges:
        r = merged_range.min_row - 1
        c = merged_range.min_col - 1
        rs = merged_range.max_row - merged_range.min_row + 1
        cs = merged_range.max_col - merged_range.min_col + 1
        key = f"{r}_{c}"
        merge_config[key] = {
            "r": r,
            "c": c,
            "rs": rs,
            "cs": cs
        }
    return merge_config


def build_column_len(sheet, max_col: int) -> Dict[str, int]:
    col_widths = {}
    for ci in range(max_col):
        col_widths[str(ci)] = get_column_width_pixels(sheet, ci)
    return col_widths


def build_row_len(sheet, max_row: int) -> Dict[str, int]:
    row_heights = {}
    for ri in range(max_row):
        h = get_row_height_pixels(sheet, ri)
        if h is not None:
            row_heights[str(ri)] = h
    return row_heights


def save_sheet_meta(sheet_val, doc_id: int, sheet_index: int, db: Session) -> ExcelSheet:
    sheet_entity = ExcelSheet(
        document_id=doc_id,
        sheet_index=sheet_index,
        sheet_name=sheet_val.title,
        active=1 if sheet_index == 0 else 0,
        status=1
    )

    max_row = sheet_val.max_row
    max_col = sheet_val.max_column
    sheet_entity.total_rows = max_row
    sheet_entity.total_cols = max_col

    # Build config elements
    merge_config = build_merge_config(sheet_val)
    column_len = build_column_len(sheet_val, max_col)
    row_len = build_row_len(sheet_val, max_row)

    sheet_entity.merge_config_json = json.dumps(merge_config)
    sheet_entity.column_len_json = json.dumps(column_len)
    sheet_entity.row_len_json = json.dumps(row_len)

    config = {
        "merge": merge_config,
        "columnlen": column_len,
        "rowlen": row_len
    }
    sheet_entity.config_json = json.dumps(config)
    sheet_entity.hyperlink_config_json = "{}"
    sheet_entity.images_config_json = "{}"
    sheet_entity.condition_format_json = "[]"
    sheet_entity.chart_json = "[]"
    sheet_entity.chunk_count = 0

    db.add(sheet_entity)
    db.flush()  # to populate sheet_entity.id

    logger.info(f"Sheet metadata saved: id={sheet_entity.id}, name={sheet_entity.sheet_name}, rows={max_row}, cols={max_col}")
    return sheet_entity


def save_sheet_chunks(sheet_val, sheet_form, sheet_entity: ExcelSheet, db: Session):
    last_row_idx = sheet_val.max_row - 1
    buffer = []
    chunk_index = 0
    chunk_start_row = 0
    chunk_end_row = 0

    for row_idx in range(last_row_idx + 1):
        # openpyxl indices are 1-based
        row_cells_val = sheet_val[row_idx + 1]
        row_cells_form = sheet_form[row_idx + 1]

        for col_idx in range(len(row_cells_val)):
            val_cell = row_cells_val[col_idx]
            form_cell = row_cells_form[col_idx]

            v_obj = build_cell_value(val_cell, form_cell)
            if v_obj is not None:
                cell_item = {
                    "r": row_idx,
                    "c": col_idx,
                    "v": v_obj
                }
                buffer.append(cell_item)

        chunk_end_row = row_idx
        is_chunk_full = (row_idx - chunk_start_row + 1) >= CHUNK_SIZE
        is_last_row = (row_idx == last_row_idx)

        if (is_chunk_full or is_last_row) and len(buffer) > 0:
            chunk = ExcelSheetChunk(
                document_id=sheet_entity.document_id,
                sheet_id=sheet_entity.id,
                chunk_index=chunk_index,
                row_start=chunk_start_row,
                row_end=chunk_end_row,
                celldata_json=json.dumps(buffer)
            )
            db.add(chunk)
            db.flush()

            buffer.clear()
            chunk_index += 1
            chunk_start_row = row_idx + 1
        elif is_chunk_full:
            chunk_start_row = row_idx + 1

    # Update sheet chunk count
    sheet_entity.chunk_count = chunk_index
    db.add(sheet_entity)
    db.flush()
    logger.info(f"Sheet chunks saved: count={chunk_index}")


def parse_and_save(file_path: str, document: ExcelDocument, db: Session) -> List[ExcelSheet]:
    """
    Parses the Excel file at file_path using openpyxl and saves metadata + chunks to the database.
    """
    logger.info(f"Starting parsing for document {document.name} at {file_path}")
    
    # Load twice to get values and formulas
    wb_val = openpyxl.load_workbook(file_path, data_only=True)
    wb_form = openpyxl.load_workbook(file_path, data_only=False)

    saved_sheets = []
    
    sheet_names = wb_val.sheetnames
    for index, sheet_name in enumerate(sheet_names):
        sheet_val = wb_val[sheet_name]
        sheet_form = wb_form[sheet_name]

        # 1. Save metadata
        sheet_entity = save_sheet_meta(sheet_val, document.id, index, db)
        saved_sheets.append(sheet_entity)

        # 2. Save cell data chunks
        save_sheet_chunks(sheet_val, sheet_form, sheet_entity, db)

    logger.info(f"Document parsing completed: {len(saved_sheets)} sheets saved")
    return saved_sheets
